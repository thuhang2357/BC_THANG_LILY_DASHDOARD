# -*- coding: utf-8 -*-
"""
Generate detail pages for dashboard
Run this after adding new monthly files
"""

import openpyxl

files = {
    '01': 'monthly_reports/THÁNG 01.26 BÁO CÁO THÁNG 01.xlsx',
    '02': 'monthly_reports/THÁNG 02.26 BÁO CÁO THÁNG 02 .xlsx',
    '03': 'monthly_reports/THÁNG 03.26 BÁO CÁO THÁNG 03.xlsx'
}

def parse_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return int(val) if val == val else 0
    s = str(val).strip().replace('.', '').replace(',', '.')
    try: return int(float(s))
    except: return 0

def fmt(v):
    return '{:,}'.format(v).replace(',', '.')

def detect_channel(first_col, second_col=''):
    """Detect channel from store name"""
    if 'STORE_MB' in first_col:
        return 'north'
    elif 'STORE_MN' in first_col:
        return 'south'
    elif 'STORE_MT' in first_col:
        return 'central'
    elif 'TMĐT' in first_col and ('BỘ PHẬN' in second_col or first_col == 'TMĐT'):
        return 'tmdt'
    elif 'VP' in first_col and 'VĂN PHÒNG' in second_col:
        return 'online'
    elif 'SEEDING' in first_col:
        return 'online'
    elif 'FB ' in first_col or 'FB-' in first_col:
        return 'online'
    elif 'IG ' in first_col or 'IG-' in first_col:
        return 'online'
    elif 'WEB' in first_col:
        return 'online'
    elif 'Đơn sỉ' in first_col:
        return 'online'
    return None

channel_names = {
    'north': 'Miền Bắc',
    'south': 'Miền Nam',
    'central': 'Miền Trung',
    'tmdt': 'TMĐT',
    'online': 'Online'
}

def generate_revenue_detail():
    """Generate revenue_detail.html from Báo cáo cửa hàng"""
    # JavaScript for filtering and totals
    filter_js = '''
    <script>
    var channelNames = {north: 'Miền Bắc', south: 'Miền Nam', central: 'Miền Trung', tmdt: 'TMĐT', online: 'Online'};
    function getChannelFromUrl() {
        var params = new URLSearchParams(window.location.search);
        return params.get('channel');
    }
    function filterByChannel() {
        var channel = getChannelFromUrl();
        var filterDiv = document.getElementById('channelFilter');
        if (channel && channelNames[channel]) {
            filterDiv.style.display = 'block';
            filterDiv.innerHTML = '<strong>Hiển thị:</strong> ' + channelNames[channel] + ' <a href="?" style="margin-left:10px">[Xem tất cả]</a>';
        }
        if (!channel) return;
        var rows = document.querySelectorAll('tbody tr');
        var totals = {dt: 0, giam_tru: 0, dt_thuan: 0, gia_von: 0, cp_ban: 0, cp_ql: 0, cp_khac: 0, lai_lo: 0};
        var hasVisible = false;
        rows.forEach(function(row) {
            var rowChannel = row.getAttribute('data-channel');
            if (rowChannel === channel) {
                row.style.display = '';
                hasVisible = true;
                // Accumulate totals
                var cells = row.querySelectorAll('.amount');
                if (cells[0]) totals.dt += parseNum(cells[0].textContent);
                if (cells[1]) totals.giam_tru += parseNum(cells[1].textContent);
                if (cells[2]) totals.dt_thuan += parseNum(cells[2].textContent);
                if (cells[3]) totals.gia_von += parseNum(cells[3].textContent);
                if (cells[4]) totals.cp_ban += parseNum(cells[4].textContent);
                if (cells[5]) totals.cp_ql += parseNum(cells[5].textContent);
                if (cells[6]) totals.cp_khac += parseNum(cells[6].textContent);
                if (cells[7]) totals.lai_lo += parseNum(cells[7].textContent);
            } else {
                row.style.display = 'none';
            }
        });
        // Update or add total row
        var tbody = document.querySelector('tbody');
        var totalRow = document.getElementById('total-row');
        if (!totalRow && hasVisible) {
            totalRow = document.createElement('tr');
            totalRow.id = 'total-row';
            totalRow.style.fontWeight = 'bold';
            totalRow.style.background = '#e8f4f8';
            totalRow.innerHTML = '<td colspan="2">TỔNG CỘNG</td><td class="amount">0</td><td class="amount">0</td><td class="amount">0</td><td class="amount">0</td><td class="amount">0</td><td class="amount">0</td><td class="amount">0</td><td class="amount">0</td><td class="amount"></td>';
            tbody.appendChild(totalRow);
        }
        if (totalRow && hasVisible) {
            var cells = totalRow.querySelectorAll('.amount');
            cells[0].textContent = fmtNum(totals.dt);
            cells[1].textContent = fmtNum(totals.giam_tru);
            cells[2].textContent = fmtNum(totals.dt_thuan);
            cells[3].textContent = fmtNum(totals.gia_von);
            cells[4].textContent = fmtNum(totals.cp_ban);
            cells[5].textContent = fmtNum(totals.cp_ql);
            cells[6].textContent = fmtNum(totals.cp_khac);
            cells[7].textContent = fmtNum(totals.lai_lo);
        }
    }
    function parseNum(s) {
        return parseInt(s.replace(/\\./g, '')) || 0;
    }
    function fmtNum(v) {
        return v.toString().replace(/\\B(?=(\\d{3})+(?!\\d))/g, '.').replace(/^/, '');
    }
    function showMonth(m) {
        document.querySelectorAll('[id^="month-"]').forEach(el => el.style.display = 'none');
        document.getElementById('month-' + m).style.display = 'block';
        document.querySelectorAll('.month-tabs button').forEach(btn => btn.classList.remove('active'));
        event.target.classList.add('active');
        filterByChannel();
    }
    window.addEventListener('DOMContentLoaded', filterByChannel);
    </script>
'''

    html = '''<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Chi tiết Doanh thu theo Cửa hàng</title>
    <style>
        body { font-family: 'Segoe UI', sans-serif; padding: 20px; background: #f5f7fa; }
        h2 { color: #1a5276; }
        h3 { color: #2980b9; margin-top: 15px; }
        table { width: 100%; border-collapse: collapse; background: white; margin-bottom: 20px; font-size: 0.85rem; }
        th, td { padding: 6px 8px; border: 1px solid #ddd; text-align: left; }
        th { background: #1a5276; color: white; }
        tr:nth-child(even) { background: #f9f9f9; }
        .amount { text-align: right; }
        .month-tabs { margin-bottom: 20px; }
        .month-tabs button { padding: 10px 20px; margin-right: 5px; cursor: pointer; border: none; border-radius: 5px; background: #ccc; }
        .month-tabs button.active { background: #1a5276; color: white; }
        .channel-filter { margin-bottom: 15px; padding: 10px; background: #e8f4f8; border-radius: 5px; display: none; }
    </style>
</head>
<body>
    <h2>Chi tiết Doanh thu (Báo cáo cửa hàng)</h2>
    <div class="channel-filter" id="channelFilter"></div>
    <div class="month-tabs">
'''
    for i, m in enumerate(files.keys()):
        cls = 'class="active"' if i == 2 else ''
        html += f'        <button onclick="showMonth(\'{m}\')" {cls}>Tháng {m}</button>\n'

    html += '    </div>\n'

    for m, f in files.items():
        try:
            wb = openpyxl.load_workbook(f, read_only=True)
            ws = wb['Báo cáo cửa hàng']
            disp = 'block' if m == '03' else 'none'
            html += f'    <div id="month-{m}" style="display:{disp}">\n'
            html += '    <table>\n    <thead><tr><th>Mã</th><th>Tên cửa hàng</th><th>Doanh thu</th><th>Giảm trừ</th><th>Doanh thu thuần</th><th>Giá vốn</th><th>CP bán hàng</th><th>CP quản lý</th><th>Chi phí khác</th><th>Lãi/Lỗ</th><th>Tỷ suất LN</th></tr></thead><tbody>\n'

            current_section = None
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row[0] and str(row[0]).strip() and not str(row[0]).startswith('='):
                    first_col = str(row[0]).strip()
                    second_col = str(row[1]).strip() if row[1] else ''

                    # Detect channel from section headers
                    new_section = detect_channel(first_col, second_col)
                    if new_section:
                        current_section = new_section

                    ma = first_col if first_col else ''
                    ten = second_col if second_col else ''
                    dt = parse_num(row[2]) if row[2] else 0
                    giam_tru = parse_num(row[3]) if row[3] else 0
                    dt_thuan = parse_num(row[4]) if row[4] else 0
                    gia_von = parse_num(row[5]) if row[5] else 0
                    cp_ban = parse_num(row[6]) if row[6] else 0
                    cp_ql = parse_num(row[7]) if row[7] else 0
                    cp_khac = parse_num(row[8]) if row[8] else 0
                    lai_lo = parse_num(row[9]) if row[9] else 0
                    ty_suat = row[10] if row[10] else ''

                    if ma and ten and (dt > 0 or dt_thuan > 0) and current_section:
                        html += f'    <tr data-channel="{current_section}"><td>{ma}</td><td>{ten}</td><td class="amount">{fmt(dt)}</td><td class="amount">{fmt(giam_tru)}</td><td class="amount">{fmt(dt_thuan)}</td><td class="amount">{fmt(gia_von)}</td><td class="amount">{fmt(cp_ban)}</td><td class="amount">{fmt(cp_ql)}</td><td class="amount">{fmt(cp_khac)}</td><td class="amount">{fmt(lai_lo)}</td><td class="amount">{ty_suat}</td></tr>\n'

            html += '    </tbody></table>\n    </div>\n'
            wb.close()
        except Exception as e:
            html += f'    <div id="month-{m}"><p>Lỗi: {str(e)}</p></div>\n'

    # Add filter JS before closing body tag - filter_js already has <script> tags inside
    html += filter_js
    html += '''
</body>
</html>
'''
    with open('dashboard/revenue_detail.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print("Created revenue_detail.html")

def generate_profit_detail():
    """Generate profit_detail.html from BCKQHĐKD"""
    html = '''<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Chi tiết Lợi nhuận</title>
    <style>
        body { font-family: 'Segoe UI', sans-serif; padding: 20px; background: #f5f7fa; }
        h2 { color: #1a5276; }
        table { width: 100%; border-collapse: collapse; background: white; margin-bottom: 20px; font-size: 0.9rem; }
        th, td { padding: 8px 10px; border: 1px solid #ddd; text-align: left; }
        th { background: #1a5276; color: white; }
        tr:nth-child(even) { background: #f9f9f9; }
        .amount { text-align: right; }
        .month-tabs { margin-bottom: 20px; }
        .month-tabs button { padding: 10px 20px; margin-right: 5px; cursor: pointer; border: none; border-radius: 5px; background: #ccc; }
        .month-tabs button.active { background: #1a5276; color: white; }
    </style>
</head>
<body>
    <h2>Chi tiết Lợi nhuận (BCKQHĐKD)</h2>
    <div class="month-tabs">
'''
    for i, m in enumerate(files.keys()):
        cls = 'class="active"' if i == 2 else ''
        html += f'        <button onclick="showMonth(\'{m}\')" {cls}>Tháng {m}</button>\n'

    html += '    </div>\n'

    for m, f in files.items():
        try:
            wb = openpyxl.load_workbook(f, read_only=True)
            ws = wb['BCKQHĐKD']
            disp = 'block' if m == '03' else 'none'
            html += f'    <div id="month-{m}" style="display:{disp}">\n'
            html += '    <table>\n    <thead><tr><th>Chỉ tiêu</th><th>Mã số</th><th>Thuyết minh</th><th>Kỳ này</th><th>Kỳ trước</th></tr></thead><tbody>\n'

            for row in ws.iter_rows(min_row=4, values_only=True):
                if row[0] and str(row[0]).strip():
                    chi_tieu = str(row[0]).strip() if row[0] else ''
                    ma_so = row[1] if row[1] else ''
                    thuyet_minh = row[2] if row[2] else ''
                    ky_nay = parse_num(row[3]) if row[3] else 0
                    ky_truoc = parse_num(row[4]) if row[4] else 0

                    if chi_tieu:
                        html += f'    <tr><td>{chi_tieu}</td><td>{ma_so}</td><td>{thuyet_minh}</td><td class="amount">{fmt(ky_nay)}</td><td class="amount">{fmt(ky_truoc)}</td></tr>\n'

            html += '    </tbody></table>\n    </div>\n'
            wb.close()
        except Exception as e:
            html += f'    <div id="month-{m}"><p>Lỗi: {str(e)}</p></div>\n'

    html += '''
    <script>
    function showMonth(m) {
        document.querySelectorAll('[id^="month-"]').forEach(el => el.style.display = 'none');
        document.getElementById('month-' + m).style.display = 'block';
        document.querySelectorAll('.month-tabs button').forEach(btn => btn.classList.remove('active'));
        event.target.classList.add('active');
    }
    </script>
</body>
</html>
'''
    with open('dashboard/profit_detail.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print("Created profit_detail.html")

def generate_expense_detail():
    """Generate expense_detail.html from BC Chi Phí"""
    html = '''<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Chi tiết Chi phí</title>
    <style>
        body { font-family: 'Segoe UI', sans-serif; padding: 20px; background: #f5f7fa; }
        h2 { color: #1a5276; }
        table { width: 100%; border-collapse: collapse; background: white; margin-bottom: 20px; font-size: 0.9rem; }
        th, td { padding: 8px 10px; border: 1px solid #ddd; text-align: left; }
        th { background: #1a5276; color: white; }
        tr:nth-child(even) { background: #f9f9f9; }
        .amount { text-align: right; }
        .month-tabs { margin-bottom: 20px; }
        .month-tabs button { padding: 10px 20px; margin-right: 5px; cursor: pointer; border: none; border-radius: 5px; background: #ccc; }
        .month-tabs button.active { background: #1a5276; color: white; }
    </style>
</head>
<body>
    <h2>Chi tiết Chi phí (BC Chi Phí)</h2>
    <div class="month-tabs">
'''
    for i, m in enumerate(files.keys()):
        cls = 'class="active"' if i == 2 else ''
        html += f'        <button onclick="showMonth(\'{m}\')" {cls}>Tháng {m}</button>\n'

    html += '    </div>\n'

    for m, f in files.items():
        try:
            wb = openpyxl.load_workbook(f, read_only=True)
            ws = wb['BC Chi Phí']
            disp = 'block' if m == '03' else 'none'
            html += f'    <div id="month-{m}" style="display:{disp}">\n'
            html += '    <table>\n    <thead><tr><th>Mã</th><th>Tên khoản mục</th><th>Kỳ trước</th><th>Kỳ này</th><th>Ghi chú</th></tr></thead><tbody>\n'

            for row in ws.iter_rows(min_row=4, values_only=True):
                if row[0] and str(row[0]).strip():
                    ma = str(row[0]).strip() if row[0] else ''
                    ten = str(row[1]).strip() if row[1] else ''
                    ky_truoc = parse_num(row[2]) if row[2] else 0
                    ky_nay = parse_num(row[3]) if row[3] else 0
                    ghichu = str(row[4]).strip() if row[4] else ''

                    if ky_nay > 0 or ky_truoc > 0:
                        html += f'    <tr><td>{ma}</td><td>{ten}</td><td class="amount">{fmt(ky_truoc)}</td><td class="amount">{fmt(ky_nay)}</td><td>{ghichu}</td></tr>\n'

            html += '    </tbody></table>\n    </div>\n'
            wb.close()
        except Exception as e:
            html += f'    <div id="month-{m}"><p>Lỗi: {str(e)}</p></div>\n'

    html += '''
    <script>
    function showMonth(m) {
        document.querySelectorAll('[id^="month-"]').forEach(el => el.style.display = 'none');
        document.getElementById('month-' + m).style.display = 'block';
        document.querySelectorAll('.month-tabs button').forEach(btn => btn.classList.remove('active'));
        event.target.classList.add('active');
    }
    </script>
</body>
</html>
'''
    with open('dashboard/expense_detail.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print("Created expense_detail.html")

def generate_all_detail_pages():
    """Generate all detail pages"""
    print("\nGenerating detail pages...")
    generate_revenue_detail()
    generate_profit_detail()
    generate_expense_detail()
    print("Done!")

if __name__ == "__main__":
    generate_all_detail_pages()
