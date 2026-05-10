import pandas as pd
import openpyxl
import json
import sys
sys.stdout.reconfigure(encoding='utf-8')

# Category groups
CATEGORY_GROUPS = {
    'nhan_su': ['TL', 'BHXH', 'CĐ', 'PĐT'],
    'thue_khau_hao': ['THUENHA', 'MMTB', 'CCDC', 'TRICHQUY', 'THUE', 'HC THUẾ'],
    'san_tmdt': ['PSS', 'PSTIKTOK', 'TUKHOA', 'NGOAISANSHOPEE'],
    'quang_cao_marketing': ['ADS - FB', 'ADS - IG', 'TIKTOK', 'VD-TIKTOK', 'TTHONG', 'KOLS', 'TCSK', 'PAGE', 'SEEDING'],
    'van_chuyen': ['SHIP', 'GHTK', 'AHA', 'SHIPHOAN'],
    'dien_nuoc_dich_vu': ['DIENNUOC', 'FPT', 'BV', 'POS', 'TĐ', 'CK', 'PCK', 'BK', 'VNPAY'],
    'chi_phi_van_hanh': ['CPCH', 'IA', 'IA (MKT)', 'VPP', 'PCT', 'BTBD', 'SPLOI', 'BB', 'TV'],
    'khac': ['CP KHÁC', 'HC', 'PQL', 'PSN', 'CATKINH', 'THEDT', 'WEB', 'SMS.', 'PM']
}

def get_sheet_name(wb):
    for name in wb.sheetnames:
        if name.strip().startswith('Chi') and ('tiết' in name or 'phi' in name):
            return name
    return None

def extract_detail_expenses(filepath, month_num):
    results = []

    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet_name = get_sheet_name(wb)
    if not sheet_name:
        wb.close()
        return results

    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[13] and row[6]:
            code = str(row[13]).strip()
            name = str(row[14]).strip() if row[14] else ''
            amount = row[6] if isinstance(row[6], (int, float)) else 0

            if amount <= 0:
                continue

            category = None
            for cat_key, codes in CATEGORY_GROUPS.items():
                if code in codes:
                    category = cat_key
                    break

            date_val = row[0]
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%d/%m/%Y')
            else:
                date_str = str(date_val) if date_val else ''

            results.append({
                'date': date_str,
                'doc': str(row[2]).strip() if row[2] else '',
                'desc': str(row[4]).strip() if row[4] else '',
                'tk': str(row[5]).strip() if row[5] else '',
                'amount': int(amount),
                'category_name': name,
                'unit': str(row[16]).strip() if row[16] else '',
                'code': code,
                'category': category,
                'month': month_num
            })

    wb.close()
    return results

# Process all 3 months
all_data = []
files = [
    ('THÁNG 01.26 BÁO CÁO THÁNG 01.xlsx', '01'),
    ('THÁNG 02.26 BÁO CÁO THÁNG 02 .xlsx', '02'),
    ('THÁNG 03.26 BÁO CÁO THÁNG 03.xlsx', '03')
]

for fname, month in files:
    print(f'Processing {fname}...')
    data = extract_detail_expenses('monthly_reports/' + fname, month)
    all_data.extend(data)
    print(f'  Month {month}: {len(data)} records')

print(f'Total: {len(all_data)} records')

# Build summary by category and code for each month
def build_summary(data):
    summary = {}
    for cat in CATEGORY_GROUPS.keys():
        summary[cat] = {'01': {}, '02': {}, '03': {}}

    for d in data:
        if d['category']:
            cat = d['category']
            month = d['month']
            code = d['code']
            name = d['category_name']

            key = code
            if key not in summary[cat][month]:
                summary[cat][month][key] = {'name': name, 'total': 0}
            summary[cat][month][key]['total'] += d['amount']

    return summary

expense_summary = build_summary(all_data)

# Output as JavaScript variable
print('\nGenerating dashboard update...')

# Create JavaScript code to embed in dashboard
js_code = 'var expenseDetailSummary = ' + json.dumps(expense_summary, ensure_ascii=False) + ';'

with open('dashboard/expense_detail_summary.js', 'w', encoding='utf-8') as f:
    f.write(js_code)

print('Done! File size:', len(js_code), 'bytes')

# Also save full data for expense_detail.html
html = '''<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Chi tiết Chi phí</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', sans-serif; padding: 20px; background: #f5f7fa; }
        h2 { color: #1a5276; margin-bottom: 15px; }
        .back-link { display: inline-block; margin-bottom: 15px; color: #1a5276; text-decoration: none; font-size: 0.9rem; }
        .back-link:hover { text-decoration: underline; }
        .month-tabs { margin-bottom: 20px; display: flex; gap: 10px; flex-wrap: wrap; }
        .month-tabs button { padding: 10px 20px; cursor: pointer; border: none; border-radius: 5px; background: #ccc; color: #333; font-size: 0.9rem; }
        .month-tabs button.active { background: #1a5276; color: white; }
        .month-tabs button:hover { background: #2980b9; color: white; }
        table { width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.1); margin-bottom: 20px; font-size: 0.85rem; }
        th, td { padding: 8px 10px; text-align: left; border-bottom: 1px solid #eee; }
        th { background: #1a5276; color: white; font-weight: 600; text-transform: uppercase; font-size: 0.75rem; }
        td.amount { text-align: right; }
        td.date { text-align: center; width: 90px; }
        td.doc { text-align: center; width: 120px; }
        tr:hover { background: #f8f9fa; }
        tr:last-child td { border-bottom: none; }
        tr.total-row { background: #e8f6f3; font-weight: 600; }
        tr.total-row td { border-top: 2px solid #1abc9c; }
        .no-data { text-align: center; padding: 40px; color: #999; }
        .category-filter { margin-bottom: 15px; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }
        .category-filter label { font-weight: 600; margin-right: 5px; }
        .category-filter select, .category-filter button { padding: 8px 12px; border: 1px solid #ddd; border-radius: 5px; font-size: 0.9rem; }
        .category-filter button { background: #1a5276; color: white; cursor: pointer; border: none; }
        .category-filter button:hover { background: #2980b9; }
        .info-text { font-size: 0.85rem; color: #666; margin-bottom: 15px; }
    </style>
</head>
<body>
    <a href="dashboard.html" class="back-link">← Quay lại Dashboard</a>
    <h2>Chi Tiết Chi Phí - Theo Nhóm</h2>

    <div class="info-text" id="info-text">Chọn nhóm chi phí để xem chi tiết từng khoản mục.</div>

    <div class="category-filter">
        <label>Lọc theo nhóm:</label>
        <select id="category-select">
            <option value="">-- Tất cả --</option>
            <option value="nhan_su">Nhân sự</option>
            <option value="thue_khau_hao">Thuê nhà & Khấu hao</option>
            <option value="san_tmdt">Sàn TMĐT</option>
            <option value="quang_cao_marketing">Quảng cáo & Marketing</option>
            <option value="van_chuyen">Vận chuyển</option>
            <option value="dien_nuoc_dich_vu">Điện nước & Dịch vụ</option>
            <option value="chi_phi_van_hanh">Chi phí vận hành</option>
            <option value="khac">Khác</option>
        </select>
        <button onclick="filterByCategory()">Lọc</button>
    </div>

    <div class="month-tabs">
        <button onclick="showMonth('01')" id="btn-01">Tháng 01</button>
        <button onclick="showMonth('02')" id="btn-02">Tháng 02</button>
        <button onclick="showMonth('03')" id="btn-03" class="active">Tháng 03</button>
    </div>

    <div id="month-01" style="display:none"></div>
    <div id="month-02" style="display:none"></div>
    <div id="month-03" style="display:block"></div>

    <script>
    var expenseData = %DATA%;

    var categoryNames = {
        'nhan_su': 'Nhân sự',
        'thue_khau_hao': 'Thuê nhà & Khấu hao',
        'san_tmdt': 'Sàn TMĐT',
        'quang_cao_marketing': 'Quảng cáo & Marketing',
        'van_chuyen': 'Vận chuyển',
        'dien_nuoc_dich_vu': 'Điện nước & Dịch vụ',
        'chi_phi_van_hanh': 'Chi phí vận hành',
        'khac': 'Khác'
    };

    function fmt(num) {
        if (num === null || num === undefined) return '0';
        return num.toString().replace(/\\B(?=(\\d{3})+(?!\\d))/g, '.');
    }

    function getUrlParam(name) {
        var params = new URLSearchParams(window.location.search);
        return params.get(name);
    }

    function filterByCategory() {
        var select = document.getElementById('category-select');
        currentCategory = select.value;
        renderAllMonths();
    }

    var currentCategory = '';
    var currentMonth = '03';

    function showMonth(m) {
        currentMonth = m;
        document.querySelectorAll('[id^="month-"]').forEach(el => el.style.display = 'none');
        document.getElementById('month-' + m).style.display = 'block';
        document.querySelectorAll('.month-tabs button').forEach(btn => btn.classList.remove('active'));
        document.getElementById('btn-' + m).classList.add('active');
    }

    function buildTable(data, month) {
        var filtered = data.filter(function(d) {
            if (d.month !== month) return false;
            if (currentCategory && d.category !== currentCategory) return false;
            return true;
        });

        if (filtered.length === 0) {
            return '<div class="no-data">Không có dữ liệu cho nhóm này</div>';
        }

        filtered.sort(function(a, b) {
            return (a.date || '').localeCompare(b.date || '');
        });

        var total = filtered.reduce(function(sum, d) { return sum + d.amount; }, 0);

        var html = '<table><thead><tr>';
        html += '<th>Ngày</th><th>Số chứng từ</th><th>Diễn giải</th><th>TK Đ/ứng</th><th>Phát sinh Nợ</th><th>Tên khoản mục</th><th>Đơn vị</th>';
        html += '</tr></thead><tbody>';

        filtered.forEach(function(d) {
            html += '<tr>';
            html += '<td class="date">' + d.date + '</td>';
            html += '<td class="doc">' + d.doc + '</td>';
            html += '<td>' + d.desc + '</td>';
            html += '<td>' + d.tk + '</td>';
            html += '<td class="amount">' + fmt(d.amount) + '</td>';
            html += '<td>' + d.category_name + '</td>';
            html += '<td>' + d.unit + '</td>';
            html += '</tr>';
        });

        html += '<tr class="total-row">';
        html += '<td colspan="4">Tổng cộng (' + filtered.length + ' dòng)</td>';
        html += '<td class="amount">' + fmt(total) + '</td>';
        html += '<td colspan="2"></td>';
        html += '</tr></tbody></table>';

        return html;
    }

    function renderAllMonths() {
        document.getElementById('month-01').innerHTML = buildTable(expenseData, '01');
        document.getElementById('month-02').innerHTML = buildTable(expenseData, '02');
        document.getElementById('month-03').innerHTML = buildTable(expenseData, '03');

        var infoText = document.getElementById('info-text');
        if (currentCategory) {
            infoText.textContent = 'Hiển thị: ' + (categoryNames[currentCategory] || currentCategory);
        } else {
            infoText.textContent = 'Hiển thị tất cả các khoản mục chi phí';
        }
    }

    function init() {
        var urlCategory = getUrlParam('category');
        if (urlCategory) {
            currentCategory = urlCategory;
            document.getElementById('category-select').value = urlCategory;
        }
        renderAllMonths();
    }

    init();
    </script>
</body>
</html>'''

html = html.replace('%DATA%', json.dumps(all_data, ensure_ascii=False))

with open('dashboard/expense_detail.html', 'w', encoding='utf-8') as f:
    f.write(html)

print('expense_detail.html generated!')