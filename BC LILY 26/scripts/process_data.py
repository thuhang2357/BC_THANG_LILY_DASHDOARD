#!/usr/bin/env python3
"""
process_data.py - ETL script to extract financial data from Excel reports
Usage: python scripts/process_data.py
"""

import pandas as pd
import json
import os
import re
import openpyxl
from pathlib import Path

# 8 major categories mapping from raw category codes (Chi tiết chi phí sheet)
CATEGORY_GROUPS = {
    'nhan_su': ['TL', 'BHXH', 'CĐ', 'PĐT'],
    'thue_khau_hao': ['THUENHA', 'MMTB', 'CCDC', 'TRICHQUY', 'THUE', 'HC THUẾ'],
    'san_tmdt': ['PSS', 'PSTIKTOK', 'TUKHOA', 'NGOAISANSHOPEE'],
    'quang_cao_marketing': ['ADS - FB', 'ADS - IG', 'TIKTOK', 'VD-TIKTOK', 'TTHONG', 'KOLS', 'TCSK', 'PAGE', 'SEEDING'],
    'van_chuyen': ['SHIP', 'GHTK', 'AHA', 'SHIPHOAN'],
    'dien_nuoc_dich_vu': ['DIENNUOC', 'FPT', 'BV', 'POS', 'TĐ', 'CK', 'PCK', 'BK', 'VNPAY'],
    'chi_phi_van_hanh': ['CPCH', 'IA', 'IA (MKT)', 'VPP', 'PCT', 'BTBD', 'SPLOI', 'BB', 'TV'],
    'khac': ['CP KHÁC', 'HC', 'PQL', 'PSN', 'CATKINH', 'THEDT', 'WEB', 'SMS.', 'PM']
}

def parse_vietnamese_number(val):
    """Parse Vietnamese number format (1.234.567) to int, handling both Excel floats and Vietnamese strings"""
    if pd.isna(val):
        return 0
    if isinstance(val, (int, float)):
        return int(round(val)) if not pd.isna(val) else 0
    val = str(val).strip()
    if not val:
        return 0
    # Check if it's a float string (e.g., '1382682085.2206335' from Excel)
    # vs Vietnamese format (e.g., '1.234.567' where dots are thousands separators)
    if '.' in val and val.count('.') == 1 and val.replace('.', '').replace('-', '').isdigit():
        # Looks like a float - parse as float then round
        try:
            return int(round(float(val)))
        except:
            pass
    # Vietnamese format: dots are thousands separators, comma is decimal
    val = val.replace('.', '').replace(',', '.').replace(' ', '')
    try:
        return int(float(val))
    except:
        return 0

def find_excel_files():
    """Find all monthly report Excel files"""
    pattern = re.compile(r'THÁNG\s+\d+\.26\s+BÁO\s+CÁO\s+THÁNG\s+\d+\s*\.xlsx', re.IGNORECASE)
    files = []
    # Check both current directory and monthly_reports folder
    search_paths = [Path('.'), Path('./monthly_reports')]
    for search_path in search_paths:
        if search_path.exists():
            for f in search_path.iterdir():
                if f.is_file() and pattern.match(f.name):
                    if not any(f.name == existing[0] for existing in files):
                        files.append((f.name, f))
    return sorted(files, key=lambda x: x[0])

def extract_summary_data(df):
    """Extract summary metrics from BCKQHĐKD sheet"""
    data = {
        'revenue': 0,
        'revenueDiscount': 0,
        'netRevenue': 0,
        'cogs': 0,
        'grossProfit': 0,
        'sellingExpense': 0,
        'managementExpense': 0,
        'netProfit': 0
    }

    for idx, row in df.iterrows():
        row_values = [str(v).strip() if pd.notna(v) else '' for v in row.values]
        first_col = row_values[0].replace(',', '.').replace(' ', '')

        if first_col.startswith('1.') and 'Doanh thu' in row_values[0] and 'bán hàng' in row_values[0]:
            data['revenue'] = parse_vietnamese_number(row_values[3])
        elif first_col.startswith('2.') and 'giảm trừ' in row_values[0]:
            data['revenueDiscount'] = parse_vietnamese_number(row_values[3])
        elif first_col.startswith('3.') and 'Doanh thu thuần' in row_values[0]:
            data['netRevenue'] = parse_vietnamese_number(row_values[3])
        elif first_col.startswith('4.') and 'Giá vốn' in row_values[0]:
            data['cogs'] = parse_vietnamese_number(row_values[3])
        elif first_col.startswith('5.') and 'Lợi nhuận gộp' in row_values[0]:
            data['grossProfit'] = parse_vietnamese_number(row_values[3])
        elif first_col.startswith('8.') and 'Chi phí bán hàng' in row_values[0]:
            data['sellingExpense'] = parse_vietnamese_number(row_values[3])
        elif first_col.startswith('9.') and 'Chi phí quản lý' in row_values[0]:
            data['managementExpense'] = parse_vietnamese_number(row_values[3])
        elif first_col.startswith('17.') and 'Lợi nhuận sau thuế' in row_values[0]:
            data['netProfit'] = parse_vietnamese_number(row_values[3])

    return data

def extract_expense_data(df):
    """Extract expense breakdown from BC Chi Phí sheet"""
    expenses = {
        # Quang cao
        'ads_fb': 0,
        'ads_ig': 0,
        'ads_tiktok': 0,
        'ads_san': 0,
        'truyen_thong': 0,
        'video_tiktok': 0,
        'gia_han_page': 0,
        # Sàn TMDT
        'san_shopee': 0,
        'san_tiktok': 0,
        'phi_noi_san': 0,
        # Van chuyen
        'ship_chemi': 0,
        'ship_ghtk': 0,
        'ship_aha': 0,
        'ship_hoan': 0,
        # Dịch vụ
        'dien_nuoc': 0,
        'internet': 0,
        'bao_ve': 0,
        'pos': 0,
        'tong_dai': 0,
        'in_an': 0,
        'chuyen_khoan': 0,
        'cong_tac': 0,
        'dao_tao': 0,
        'tin_nhan': 0,
        'thi_cong': 0,
        'sang_nhuong': 0,
        'tu_van': 0,
        # Luong
        'tien_luong': 0,
        'bhxh': 0,
        'cong_doan': 0,
        # Phan bo
        'phan_bo_chung': 0,
        'phan_bo_ccdc': 0,
        # May moc
        'may_moc': 0,
        'phan_mem': 0,
        'website': 0,
        # Thue nha
        'thue_nha': 0,
        'thue_bang': 0,
        # Trich quy
        'trich_quy': 0,
        # Marketing
        'mkt_thue_ngoai': 0,
        'kols': 0,
        'su_kien': 0,
        'seeding': 0,
        # Khac
        'khac': 0,
        'cp_khac': 0,
        'chi_phi_cua_hang': 0,
        'bao_bi': 0,
        'bao_tri': 0,
        'san_pham_loi': 0,
        'van_phong_pham': 0,
        'thue': 0,
        'bhyt': 0,
        'hanh_chinh': 0,
        'dien_thoai': 0,
        'chuyen_tien': 0
    }

    for idx, row in df.iterrows():
        if idx < 4:
            continue
        row_values = [str(v).strip() if pd.notna(v) else '' for v in row.values]
        if len(row_values) < 4:
            continue

        name = row_values[1].lower() if len(row_values) > 1 else ''
        name_orig = row_values[1] if len(row_values) > 1 else ''
        amount = parse_vietnamese_number(row_values[3])

        if amount <= 0:
            continue

        # Ads Facebook
        if 'ads fb' in name or 'chạy ads fb' in name:
            expenses['ads_fb'] += amount
        # Ads Instagram
        elif 'ads ig' in name or 'chạy ads ig' in name:
            expenses['ads_ig'] += amount
        # Ads TikTok
        elif 'ads tiktok' in name or 'chạy ads tiktok' in name:
            expenses['ads_tiktok'] += amount
        # Video TikTok
        elif 'dựng, cs video tiktok' in name or 'video tiktok' in name:
            expenses['video_tiktok'] += amount
        # Gia han page
        elif 'gia hạn page' in name:
            expenses['gia_han_page'] += amount
        # Truyen thong
        elif 'truyền thông' in name:
            expenses['truyen_thong'] += amount
        # Ads san / quang cao
        elif any(x in name for x in ['quảng cáo', 'mua gói quảng cáo']):
            expenses['ads_san'] += amount
        # San Shopee
        elif 'sàn shopee' in name or 'chi phí sàn shopee' in name:
            expenses['san_shopee'] += amount
        # San TikTok
        elif 'sàn tiktok' in name or 'phí sàn tiktok' in name:
            expenses['san_tiktok'] += amount
        # Phi noi san
        elif 'phí nội sàn' in name:
            expenses['phi_noi_san'] += amount
        # Ship Chemi
        elif 'ship chemi' in name:
            expenses['ship_chemi'] += amount
        # Ship GHTK
        elif 'ghtk' in name:
            expenses['ship_ghtk'] += amount
        # Ship AHA
        elif 'aha' in name:
            expenses['ship_aha'] += amount
        # Ship hoan
        elif 'ship hoàn' in name:
            expenses['ship_hoan'] += amount
        # Ship
        elif 'ship' in name:
            expenses['ship_chemi'] += amount
        # Dien nuoc
        elif 'điện' in name or 'nước' in name:
            expenses['dien_nuoc'] += amount
        # Internet
        elif 'internet' in name:
            expenses['internet'] += amount
        # Bao ve
        elif 'bảo vệ' in name:
            expenses['bao_ve'] += amount
        # POS
        elif 'pos' in name:
            expenses['pos'] += amount
        # Tong dai
        elif 'tổng đài' in name:
            expenses['tong_dai'] += amount
        # In an
        elif 'in ấn' in name or 'in an' in name:
            expenses['in_an'] += amount
        # Chuyen khoan
        elif 'chuyển khoản' in name or 'chuyen khoan' in name:
            expenses['chuyen_khoan'] += amount
        # Cong tac
        elif 'công tác' in name or 'cong tac' in name:
            expenses['cong_tac'] += amount
        # Dao tao
        elif 'đào tạo' in name or 'đào tạo' in name:
            expenses['dao_tao'] += amount
        # Tin nhan
        elif 'tin nhắn' in name or 'tin nhan' in name:
            expenses['tin_nhan'] += amount
        # Thi cong
        elif 'thi công' in name or 'thi cong' in name:
            expenses['thi_cong'] += amount
        # Sang nhượng
        elif 'sang nhượng' in name or 'sang nhuong' in name:
            expenses['sang_nhuong'] += amount
        # Tu van
        elif 'tư vấn' in name or 'tu van' in name:
            expenses['tu_van'] += amount
        # Dịch vụ khác
        elif 'dịch vụ' in name or 'dich vu' in name:
            expenses['dien_nuoc'] += amount
        # Tien luong
        elif 'tiền lương' in name or 'tien luong' in name:
            expenses['tien_luong'] += amount
        # BHXH
        elif 'bảo hiểm xã hội' in name or 'bhxh' in name:
            expenses['bhxh'] += amount
        # Cong doan
        elif 'công đoàn' in name or 'cong doan' in name:
            expenses['cong_doan'] += amount
        # Phan bo chung
        elif 'phân bổ chung' in name or 'phan bo chung' in name:
            expenses['phan_bo_chung'] += amount
        # Phan bo CCDC
        elif 'phân bổ ccdc' in name or 'phan bo ccdc' in name:
            expenses['phan_bo_ccdc'] += amount
        # May moc
        elif 'máy móc' in name or 'may moc' in name:
            expenses['may_moc'] += amount
        # Phan mem
        elif 'phần mềm' in name or 'phan mem' in name:
            expenses['phan_mem'] += amount
        # Website
        elif 'website' in name:
            expenses['website'] += amount
        # Thue nha
        elif 'thuê nhà' in name or 'thue nha' in name:
            expenses['thue_nha'] += amount
        # Thue bang
        elif 'thuê bằng' in name or 'thue bang' in name:
            expenses['thue_bang'] += amount
        # Trich quy
        elif 'trích quỹ' in name or 'trich quy' in name:
            expenses['trich_quy'] += amount
        # MKT thue ngoai
        elif 'mkt thuê ngoài' in name or 'mkt thue ngoai' in name:
            expenses['mkt_thue_ngoai'] += amount
        # KOLs
        elif 'kols' in name or 'kol' in name or 'book kols' in name:
            expenses['kols'] += amount
        # Su kien
        elif 'tổ chức sự kiện' in name or 'su kien' in name:
            expenses['su_kien'] += amount
        # Seeding
        elif 'seeding' in name:
            expenses['seeding'] += amount
        # Khac
        elif 'khác' in name or 'cp khác' in name:
            expenses['khac'] += amount
        # Chi phi cua hang
        elif 'chi phí cửa hàng' in name or 'chi phi cua hang' in name:
            expenses['chi_phi_cua_hang'] += amount
        # Bao bi
        elif 'bao bì' in name or 'bao bi' in name:
            expenses['bao_bi'] += amount
        # Bao tri
        elif 'bảo trì' in name or 'bao tri' in name:
            expenses['bao_tri'] += amount
        # San pham loi
        elif 'sản phẩm lỗi' in name or 'san pham loi' in name:
            expenses['san_pham_loi'] += amount
        # Van phong pham
        elif 'văn phòng phẩm' in name or 'van phong pham' in name:
            expenses['van_phong_pham'] += amount
        # Thue
        elif 'thuế' in name and 'gtgt' not in name and 'hnh' not in name:
            expenses['thue'] += amount
        # BHYT
        elif 'bhyt' in name:
            expenses['bhyt'] += amount
        # Hanh chinh
        elif 'hành chính' in name or 'hanh chinh' in name:
            expenses['hanh_chinh'] += amount
        # Dien thoai
        elif 'điện thoại' in name or 'dien thoai' in name or 'thẻ điện' in name:
            expenses['dien_thoai'] += amount
        # Chuyen tien
        elif 'chuyển tiền' in name or 'chuyen tien' in name:
            expenses['chuyen_tien'] += amount
        else:
            expenses['khac'] += amount

    return expenses

def extract_expense_from_detail(filepath):
    """Extract expense data from Chi tiết chi phí sheet and group into 8 categories"""
    expenses = {cat: 0 for cat in CATEGORY_GROUPS.keys()}

    wb = openpyxl.load_workbook(filepath, read_only=True)
    # Find the chi tiet chi phi sheet
    ws = None
    for name in wb.sheetnames:
        if 'Chi tiết chi phí' in name or 'Chi phí chi tiết' in name:
            ws = wb[name]
            break

    if ws is None:
        wb.close()
        return expenses

    # Process rows (start from row 5 based on earlier analysis)
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[13] and row[6]:  # column 14 (code) and column 7 (debit amount)
            code = str(row[13]).strip()
            amount = row[6] if isinstance(row[6], (int, float)) else 0

            if amount <= 0:
                continue

            # Find which group this code belongs to
            for cat_key, codes in CATEGORY_GROUPS.items():
                if code in codes:
                    expenses[cat_key] += amount
                    break

    wb.close()
    return expenses

def extract_channel_data_from_file(filepath):
    """Extract full channel metrics from Báo cáo cửa hàng sheet using openpyxl with data_only=True"""
    channels = {
        'north': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'south': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'central': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'tmdt': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'online': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0}
    }

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if 'Báo cáo cửa hàng' not in wb.sheetnames:
        wb.close()
        return channels

    ws = wb['Báo cáo cửa hàng']

    for row in ws.iter_rows(min_row=5, values_only=True):
        first_col = str(row[0]).strip() if row[0] else ''
        second_col = str(row[1]).strip() if row[1] else ''

        if not first_col:
            continue

        channel = None
        if 'STORE_MB' in first_col:
            channel = 'north'
        elif 'STORE_MN' in first_col:
            channel = 'south'
        elif 'STORE_MT' in first_col:
            channel = 'central'
        elif 'TMĐT' in first_col and ('BỘ PHẬN' in second_col or first_col == 'TMĐT'):
            channel = 'tmdt'
        elif 'VP' in first_col and 'VĂN PHÒNG' in second_col:
            channel = 'online'

        if channel:
            # Columns: 2=revenue, 4=netRevenue, 5=cogs, 6=sellExp, 7=mgmtExp, 9=netProfit
            rev = row[2] if isinstance(row[2], (int, float)) else 0
            net_rev = row[4] if isinstance(row[4], (int, float)) else 0
            cogs = row[5] if isinstance(row[5], (int, float)) else 0
            sell_exp = row[6] if isinstance(row[6], (int, float)) else 0
            mgmt_exp = row[7] if isinstance(row[7], (int, float)) else 0
            net_profit = row[9] if isinstance(row[9], (int, float)) else 0

            channels[channel]['revenue'] = int(round(rev)) if rev else 0
            channels[channel]['netRevenue'] = int(round(net_rev)) if net_rev else 0
            channels[channel]['cogs'] = int(round(cogs)) if cogs else 0
            channels[channel]['sellingExpense'] = int(round(sell_exp)) if sell_exp else 0
            channels[channel]['managementExpense'] = int(round(mgmt_exp)) if mgmt_exp else 0
            channels[channel]['netProfit'] = int(round(net_profit)) if net_profit else 0

    wb.close()

    # Calculate derived metrics
    for ch in ['north', 'south', 'central', 'tmdt', 'online']:
        channels[ch]['grossProfit'] = channels[ch]['netRevenue'] - channels[ch]['cogs']
        channels[ch]['grossMargin'] = (channels[ch]['grossProfit'] / channels[ch]['netRevenue'] * 100) if channels[ch]['netRevenue'] > 0 else 0
        channels[ch]['totalExpense'] = channels[ch]['sellingExpense'] + channels[ch]['managementExpense']
        channels[ch]['expenseRatio'] = (channels[ch]['totalExpense'] / channels[ch]['netRevenue'] * 100) if channels[ch]['netRevenue'] > 0 else 0
        channels[ch]['netMargin'] = (channels[ch]['netProfit'] / channels[ch]['netRevenue'] * 100) if channels[ch]['netRevenue'] > 0 else 0

    return channels

def extract_channel_data(df):
    """Legacy function - use extract_channel_data_from_file instead"""
    return {
        'north': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'south': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'central': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'tmdt': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0},
        'online': {'revenue': 0, 'netRevenue': 0, 'cogs': 0, 'sellingExpense': 0, 'managementExpense': 0, 'netProfit': 0}
    }
    # Calculate derived metrics
    for ch in ['north', 'south', 'central', 'tmdt', 'online']:
        channels[ch]['grossProfit'] = channels[ch]['netRevenue'] - channels[ch]['cogs']
        channels[ch]['grossMargin'] = (channels[ch]['grossProfit'] / channels[ch]['netRevenue'] * 100) if channels[ch]['netRevenue'] > 0 else 0
        channels[ch]['totalExpense'] = channels[ch]['sellingExpense'] + channels[ch]['managementExpense']
        channels[ch]['expenseRatio'] = (channels[ch]['totalExpense'] / channels[ch]['netRevenue'] * 100) if channels[ch]['netRevenue'] > 0 else 0
        channels[ch]['netMargin'] = (channels[ch]['netProfit'] / channels[ch]['netRevenue'] * 100) if channels[ch]['netRevenue'] > 0 else 0

    return channels

def process_all_files():
    """Main ETL process"""
    files = find_excel_files()
    print(f"Found {len(files)} Excel files")

    result = {
        'meta': {
            'company': 'Công Ty Cổ Phần ULTD Thịnh Phát',
            'lastUpdated': pd.Timestamp.now().strftime('%Y-%m-%d'),
            'currency': 'VND'
        },
        'months': [],
        'monthKeys': [],
        'selectedMonths': [0, 1, 2],
        'summary': {
            'revenue': [], 'revenueDiscount': [], 'netRevenue': [],
            'cogs': [], 'grossProfit': [], 'sellingExpense': [],
            'managementExpense': [], 'netProfit': [], 'cash': []
        },
        'growth': {
            'revenue': [None], 'grossProfit': [None], 'netProfit': [None]
        },
        'expenses': {
            'nhan_su': [], 'thue_khau_hao': [], 'san_tmdt': [], 'quang_cao_marketing': [],
            'van_chuyen': [], 'dien_nuoc_dich_vu': [], 'chi_phi_van_hanh': [], 'khac': []
        },
        'channels': {
            'north': {
                'revenue': [], 'netRevenue': [], 'cogs': [],
                'grossProfit': [], 'grossMargin': [],
                'sellingExpense': [], 'managementExpense': [], 'totalExpense': [],
                'expenseRatio': [], 'netProfit': [], 'netMargin': []
            },
            'south': {
                'revenue': [], 'netRevenue': [], 'cogs': [],
                'grossProfit': [], 'grossMargin': [],
                'sellingExpense': [], 'managementExpense': [], 'totalExpense': [],
                'expenseRatio': [], 'netProfit': [], 'netMargin': []
            },
            'central': {
                'revenue': [], 'netRevenue': [], 'cogs': [],
                'grossProfit': [], 'grossMargin': [],
                'sellingExpense': [], 'managementExpense': [], 'totalExpense': [],
                'expenseRatio': [], 'netProfit': [], 'netMargin': []
            },
            'tmdt': {
                'revenue': [], 'netRevenue': [], 'cogs': [],
                'grossProfit': [], 'grossMargin': [],
                'sellingExpense': [], 'managementExpense': [], 'totalExpense': [],
                'expenseRatio': [], 'netProfit': [], 'netMargin': []
            },
            'online': {
                'revenue': [], 'netRevenue': [], 'cogs': [],
                'grossProfit': [], 'grossMargin': [],
                'sellingExpense': [], 'managementExpense': [], 'totalExpense': [],
                'expenseRatio': [], 'netProfit': [], 'netMargin': []
            }
        }
    }

    for filename, filepath in files:
        print(f"Processing: {filename}")
        month_key = re.search(r'THÁNG\s+(\d+)', filename, re.IGNORECASE).group(1)
        result['months'].append(f'Tháng {month_key}')
        result['monthKeys'].append(month_key.zfill(2))

        filepath_str = str(filepath)
        xls = pd.ExcelFile(filepath_str, engine='openpyxl')

        # Summary
        df_summary = pd.read_excel(filepath_str, sheet_name='BCKQHĐKD', header=None, engine='openpyxl')
        summary = extract_summary_data(df_summary)
        for k, v in summary.items():
            result['summary'][k].append(v)

        # Expenses - from Chi tiết chi phí sheet (8 categories)
        expenses = extract_expense_from_detail(filepath_str)
        for k, v in expenses.items():
            result['expenses'][k].append(v)

        # Channels - use openpyxl with data_only=True to get calculated values
        ch_data = extract_channel_data_from_file(filepath_str)
        for ch in ['north', 'south', 'central', 'tmdt', 'online']:
                result['channels'][ch]['revenue'].append(ch_data[ch]['revenue'])
                result['channels'][ch]['netRevenue'].append(ch_data[ch]['netRevenue'])
                result['channels'][ch]['cogs'].append(ch_data[ch]['cogs'])
                result['channels'][ch]['grossProfit'].append(ch_data[ch]['grossProfit'])
                result['channels'][ch]['grossMargin'].append(ch_data[ch]['grossMargin'])
                result['channels'][ch]['sellingExpense'].append(ch_data[ch]['sellingExpense'])
                result['channels'][ch]['managementExpense'].append(ch_data[ch]['managementExpense'])
                result['channels'][ch]['totalExpense'].append(ch_data[ch]['totalExpense'])
                result['channels'][ch]['expenseRatio'].append(ch_data[ch]['expenseRatio'])
                result['channels'][ch]['netProfit'].append(ch_data[ch]['netProfit'])
                result['channels'][ch]['netMargin'].append(ch_data[ch]['netMargin'])

    # Calculate growth
    for i in range(1, len(result['summary']['revenue'])):
        prev_rev = result['summary']['revenue'][i-1]
        curr_rev = result['summary']['revenue'][i]
        if prev_rev > 0:
            result['growth']['revenue'].append(round((curr_rev - prev_rev) / prev_rev * 100, 1))
        else:
            result['growth']['revenue'].append(None)

        prev_profit = result['summary']['grossProfit'][i-1]
        curr_profit = result['summary']['grossProfit'][i]
        if prev_profit > 0:
            result['growth']['grossProfit'].append(round((curr_profit - prev_profit) / prev_profit * 100, 1))
        else:
            result['growth']['grossProfit'].append(None)

        prev_net = result['summary']['netProfit'][i-1]
        curr_net = result['summary']['netProfit'][i]
        if prev_net != 0:
            result['growth']['netProfit'].append(round((curr_net - prev_net) / abs(prev_net) * 100, 1))
        else:
            result['growth']['netProfit'].append(None)

    # Save to data.json
    with open('data/data.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"Data saved to data/data.json")
    print(f"Summary: {len(result['months'])} months processed")
    return result

if __name__ == '__main__':
    process_all_files()