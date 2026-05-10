# -*- coding: utf-8 -*-
"""
Extract expense data from "Chi tiết chi phí" sheet
Groups 53 categories into 8 major categories
"""

import openpyxl
import sys
import io

# Fix encoding for Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 8 major categories mapping from raw category codes
CATEGORY_GROUPS = {
    'nhan_su': ['TL', 'BHXH', 'CĐ', 'PĐT'],
    'thue_khau_hao': ['THUENHA', 'MMTB', 'CCDC', 'TRICHQUY', 'THUE', 'HC THUẾ'],
    'san_tmdt': ['PSS', 'PSTIKTOK', 'TUKHOA', 'NGOAISANSHOPEE'],
    'quang_cao_marketing': ['ADS - FB', 'ADS - IG', 'TIKTOK', 'VD-TIKTOK', 'TTHONG', 'KOLS', 'TCSK', 'PAGE', 'SEEDING'],
    'van_chuyen': ['SHIP', 'GHTK', 'AHA', 'SHIPHOAN'],
    'dien_nuoc_dich_vu': ['DIENNUOC', 'FPT', 'BV', 'POS', 'TĐ', 'CK', 'PCK', 'BK', 'VNPAY'],
    'chi_phi_van_hanh': ['CPCH', 'IA', 'IA (MKT)', 'VPP', 'PCT', 'BTBD', 'SPLOI', 'BB', 'TV'],
    'khac': ['CP KHÁC', 'HC', 'PQL', 'PSN', 'CATKINH', 'THEDT', 'WEB', 'SMS.', 'PM']
}

CATEGORY_TITLES = {
    'nhan_su': 'Nhân sự',
    'thue_khau_hao': 'Thuê nhà & Khấu hao',
    'san_tmdt': 'Sàn TMĐT',
    'quang_cao_marketing': 'Quảng cáo & Marketing',
    'van_chuyen': 'Vận chuyển',
    'dien_nuoc_dich_vu': 'Điện nước & Dịch vụ',
    'chi_phi_van_hanh': 'Chi phí vận hành',
    'khac': 'Khác'
}

def extract_expense_detail():
    """Extract expense data from Chi tiết chi phí sheet"""
    files = {
        '01': 'monthly_reports/THÁNG 01.26 BÁO CÁO THÁNG 01.xlsx',
        '02': 'monthly_reports/THÁNG 02.26 BÁO CÁO THÁNG 02 .xlsx',
        '03': 'monthly_reports/THÁNG 03.26 BÁO CÁO THÁNG 03.xlsx'
    }

    results = {}
    for month, filepath in files.items():
        wb = openpyxl.load_workbook(filepath, read_only=True)
        # Find the chi tiet chi phi sheet
        ws = None
        for name in wb.sheetnames:
            if 'Chi tiết chi phí' in name or 'Chi phí chi tiết' in name:
                ws = wb[name]
                break

        if ws is None:
            print(f"Warning: Could not find Chi tiet chi phi sheet in {filepath}")
            wb.close()
            continue

        # Initialize category totals
        category_totals = {cat: 0 for cat in CATEGORY_GROUPS.keys()}

        # Process rows (start from row 5 based on earlier analysis)
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[13] and row[6]:  # column 14 (code) and column 7 (debit amount)
                code = str(row[13]).strip()
                amount = row[6] if isinstance(row[6], (int, float)) else 0

                if amount <= 0:
                    continue

                # Find which group this code belongs to
                for cat_key, codes in CATEGORY_GROUPS.items():
                    if code in codes:
                        category_totals[cat_key] += amount
                        break

        results[month] = category_totals
        wb.close()

    return results

def get_expense_summary():
    """Get formatted expense summary for all 3 months"""
    data = extract_expense_detail()

    # Convert to array format [month1, month2, month3] for each category
    expense_summary = {}
    for cat in CATEGORY_GROUPS.keys():
        expense_summary[cat] = [
            data.get('01', {}).get(cat, 0),
            data.get('02', {}).get(cat, 0),
            data.get('03', {}).get(cat, 0)
        ]

    return expense_summary, CATEGORY_TITLES

if __name__ == "__main__":
    expense_summary, titles = get_expense_summary()
    print("Expense Summary (8 categories):")
    for cat, values in expense_summary.items():
        print(f"  {titles[cat]}: {values}")
